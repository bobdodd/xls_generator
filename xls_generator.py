import pandas as pd
from jinja2 import Environment, FileSystemLoader
from datetime import datetime
from pymongo import MongoClient
import openpyxl
from openpyxl.styles import Alignment
from urllib.parse import urlparse
from collections import defaultdict
import json

class TemplateAnalyzer:
    def __init__(self, db):
        self.db = db
        self.structures = defaultdict(set)
        self.examples = {}

    def analyze_test_structures(self):
        """Analyze all test structures in the database"""
        results = self.db.page_results.find()
        
        for result in results:
            if 'results' in result and 'accessibility' in result['results']:
                tests = result['results']['accessibility'].get('tests', {})
                
                # Analyze each test type
                for test_name, test_data in tests.items():
                    if isinstance(test_data, dict):
                        self._record_structure(self.structures[test_name], test_data)
                        
                        # Store first example if we don't have one yet
                        if test_name not in self.examples:
                            self.examples[test_name] = test_data
        
        return self.structures, self.examples

    def _record_structure(self, structure_set, data, path=""):
        """Recursively record the structure of a dictionary"""
        if isinstance(data, dict):
            for key, value in data.items():
                new_path = f"{path}.{key}" if path else key
                structure_set.add(f"{new_path} ({type(value).__name__})")
                self._record_structure(structure_set, value, new_path)
        elif isinstance(data, list) and data:
            # Record structure of first item in list as example
            if data and isinstance(data[0], dict):
                self._record_structure(structure_set, data[0], f"{path}[]")

    def print_analysis(self):
        """Print a summary of all test types and their structures"""
        print("\nTest Types Analysis Summary:")
        print("=" * 50)
        
        for test_name in sorted(self.structures.keys()):
            print(f"\n{test_name}:")
            print("-" * len(test_name))
            for path in sorted(self.structures[test_name]):
                print(f"  {path}")

DEFAULT_DB_NAME = 'accessibility_tests'

class AccessibilityDB:
    def __init__(self, db_name=None):
        try:
            self.client = MongoClient('mongodb://localhost:27017/')
            
            # Use the specified database name or default
            if db_name is None:
                db_name = DEFAULT_DB_NAME
                print(f"Warning: No database name specified for XLS generator. Using default database '{DEFAULT_DB_NAME}'.")
            
            self.db_name = db_name
            self.db = self.client[db_name]
            self.test_runs = self.db['test_runs']
            self.page_results = self.db['page_results']
            
            print(f"XLS Generator connected to database: '{db_name}'")
        except Exception as e:
            print(f"Failed to connect to MongoDB: {e}")
            raise

    def get_page_results(self, test_run_ids=None):
        """Get all page results for specific test runs"""
        query = {}
        if test_run_ids:
            if isinstance(test_run_ids, list):
                query['test_run_id'] = {'$in': test_run_ids}
            else:
                query['test_run_id'] = test_run_ids
        return list(self.page_results.find(query))

    def get_most_recent_test_run_id(self):
        """Get the most recent test run ID"""
        latest_run = self.test_runs.find_one(
            sort=[('timestamp_start', -1)]
        )
        return str(latest_run['_id']) if latest_run else None

    def get_all_test_run_ids(self):
        """Get all test run IDs"""
        test_runs = self.test_runs.find({}, {'_id': 1})
        return [str(run['_id']) for run in test_runs]
        
    def get_test_run_by_name(self, name):
        """Get a test run by name"""
        test_run = self.test_runs.find_one({"name": name})
        return test_run

    def __del__(self):
        if hasattr(self, 'client'):
            self.client.close()

class AccessibilityReportGenerator:
    def __init__(self, db, structures=None, examples=None):
        self.db = db
        self.env = Environment(loader=FileSystemLoader('templates'))
        self.structures = structures or {}
        self.examples = examples or {}
        self.test_documentation = {}  # Store test documentation from all tests

    def collect_test_documentation(self, results):
        """
        Extract and collect test documentation from test results and test run metadata
        
        Args:
            results: List of page result documents from MongoDB
            
        Returns:
            Dictionary with test documentation organized by test type
        """
        total_docs_found = 0
        
        # First, check if documentation exists in the test_runs collection
        print("Checking test_runs collection for documentation...")
        try:
            # Get all test runs
            test_runs = list(self.db.test_runs.find({}))
            for test_run in test_runs:
                # Check if this test run has documentation
                if 'documentation' in test_run:
                    documentation = test_run.get('documentation', {})
                    if isinstance(documentation, dict):
                        for test_name, doc_data in documentation.items():
                            if test_name not in self.test_documentation:
                                print(f"Found documentation for {test_name} in test_run {test_run.get('_id')}")
                                self.test_documentation[test_name] = doc_data
                                total_docs_found += 1
                
                # Also check for a nested tests collection within the test run
                if 'tests' in test_run:
                    tests = test_run.get('tests', {})
                    for test_name, test_data in tests.items():
                        if isinstance(test_data, dict) and 'documentation' in test_data:
                            if test_name not in self.test_documentation:
                                print(f"Found documentation for {test_name} in test_run.tests")
                                self.test_documentation[test_name] = test_data['documentation']
                                total_docs_found += 1
        except Exception as e:
            print(f"Error checking test_runs collection: {str(e)}")
            
        # Then continue with checking page results
        for result in results:
            print(f"Processing result for URL: {result.get('url', 'unknown')}")
            
            if 'results' in result and 'accessibility' in result['results']:
                tests = result['results']['accessibility'].get('tests', {})
                print(f"Found {len(tests)} test types in this result")
                
                for test_name, test_data in tests.items():
                    print(f"  Checking {test_name} test data...")
                    
                    if isinstance(test_data, dict):
                        # Check for documentation directly in the test result data (new structure)
                        print(f"  Keys in {test_name} test data: {list(test_data.keys())}")
                        if 'documentation' in test_data:
                            print(f"  Found documentation directly in {test_name} -> documentation")
                            doc_data = test_data['documentation']
                            print(f"  Documentation content: {doc_data.get('testName', 'N/A')}, tests: {len(doc_data.get('tests', []))}")
                            if test_name in self.test_documentation:
                                print(f"  Already have documentation for {test_name}, skipping duplicate")
                            else:
                                self.test_documentation[test_name] = test_data['documentation']
                                total_docs_found += 1
                                print(f"  Added documentation for {test_name} with {len(test_data['documentation'].get('tests', []))} individual tests")
                            continue
                            
                        # First, check for documentation in nested test_data if it's nested
                        if test_name in test_data and 'documentation' in test_data[test_name]:
                            print(f"  Found nested documentation in {test_name} -> {test_name} -> documentation")
                            if test_name in self.test_documentation:
                                print(f"  Already have documentation for {test_name}, skipping duplicate")
                            else:
                                self.test_documentation[test_name] = test_data[test_name]['documentation']
                                total_docs_found += 1
                                print(f"  Added documentation for {test_name} with {len(test_data[test_name]['documentation'].get('tests', []))} individual tests")
                        else:
                            # Extra debugging for example.com tests
                            if result.get('url') == 'https://example.com':
                                print(f"  No documentation found in {test_name} test data. Keys at this level: {list(test_data.keys())}")
                                # If there's a nested structure with the same name, inspect it
                                if test_name in test_data:
                                    print(f"  Found nested {test_name} object. Keys: {list(test_data[test_name].keys())}")
                                    if 'documentation' in test_data[test_name]:
                                        print(f"  Found documentation in nested structure!")
                                        if test_name in self.test_documentation:
                                            print(f"  Already have documentation for {test_name}, skipping duplicate")
                                        else:
                                            self.test_documentation[test_name] = test_data[test_name]['documentation']
                                            total_docs_found += 1
                                            print(f"  Added documentation for {test_name} with {len(test_data[test_name]['documentation'].get('tests', []))} individual tests")
        
        # Add a separate step to directly check for documentation in specific locations
        # This addresses the different ways tests might structure their output
        print("\nAttempting comprehensive documentation scan with multiple strategies...")
        
        # Map of possible test output field names based on test names
        test_output_fields = {
            'images': 'images',
            'tables': 'tables', 
            'headings': 'headings',
            'focus_management': 'focus_management',
            'page_structure': 'page_structure',
            'accessible_names': 'accessible_names',
            'landmarks': 'landmarks',
            'forms': 'forms',
            'colors': 'colors',
            'html_structure': 'html_structure',
            'animations': 'animations',
            'videos': 'videos'
            # Add more mappings as needed
        }
        
        print("Scanning database for test documentation using multiple approaches...")
        for result in results:
            if 'results' in result and 'accessibility' in result['results']:
                tests = result['results']['accessibility'].get('tests', {})
                
                for test_name, test_data in tests.items():
                    # Skip if we already found documentation for this test
                    if test_name in self.test_documentation:
                        continue
                    
                    # 1. Check if the test itself contains documentation field
                    if isinstance(test_data, dict) and 'documentation' in test_data:
                        print(f"Found direct documentation in test data for {test_name}")
                        self.test_documentation[test_name] = test_data['documentation']
                        total_docs_found += 1
                        continue
                    
                    # 2. Check if test has a corresponding field with documentation
                    output_field = test_output_fields.get(test_name, test_name)
                    if isinstance(test_data, dict) and output_field in test_data:
                        output_data = test_data[output_field]
                        if isinstance(output_data, dict) and 'documentation' in output_data:
                            print(f"Found documentation in {test_name} -> {output_field} field")
                            self.test_documentation[test_name] = output_data['documentation']
                            total_docs_found += 1
                            continue
                    
                    # 3. For specific multi-word test names, try direct matching of test field
                    if '_' in test_name:
                        # Try with underscores
                        if test_name in test_data:
                            field_data = test_data[test_name]
                            if isinstance(field_data, dict) and 'documentation' in field_data:
                                print(f"Found documentation using underscore name match: {test_name}")
                                self.test_documentation[test_name] = field_data['documentation']
                                total_docs_found += 1
                                continue
                    
                    # 4. Look inside the main result for top-level fields that might contain documentation
                    # This checks for potential custom field names not in our mapping
                    for field_name, field_data in test_data.items():
                        if isinstance(field_data, dict) and 'documentation' in field_data:
                            print(f"Found documentation in {test_name} -> {field_name} field")
                            self.test_documentation[test_name] = field_data['documentation']
                            total_docs_found += 1
                            break
        
        print(f"Collected documentation for {len(self.test_documentation)} test types (found {total_docs_found} new)")
        return self.test_documentation
    
    def format_issue_name(self, test_name, flag_name):
        """Format issue name consistently with documentation if available"""
        # Remove 'has' prefix
        flag_text = flag_name[3:] if flag_name.startswith('has') else flag_name
        
        # Split camel case into words
        words = []
        current_word = ''
        for char in flag_text:
            if char.isupper() and current_word:
                words.append(current_word)
                current_word = char
            else:
                current_word += char
        words.append(current_word)
        
        # Format issue type and test name
        issue_type = ' '.join(words).title()
        formatted_test_name = ' '.join(
            word.capitalize() 
            for word in test_name.split('_')
        )
        
        # Try multiple variants of the test name to match documentation
        test_name_variants = [
            test_name,                     # Original name
            test_name.replace('-', '_'),   # Replace hyphens with underscores
            test_name.replace('_', '-'),   # Replace underscores with hyphens
            test_name.lower(),             # Lowercase
            test_name.replace('-', '').replace('_', '')  # No separators
        ]
        
        # Check if we have documentation for any variant of this test name
        for variant in test_name_variants:
            if variant in self.test_documentation:
                docs = self.test_documentation[variant]
                doc_test_name = docs.get('testName', formatted_test_name)
                
                # Try to find a specific test that matches this flag
                for test in docs.get('tests', []):
                    results_fields = test.get('resultsFields', {})
                    
                    # Look for various ways the flag could be referenced
                    flag_variants = [
                        f"pageFlags.{flag_name}",
                        f"pageFlags.has{flag_text}",
                        f"details.{flag_name}",
                        flag_name
                    ]
                    
                    # Look for any matching result field
                    for field in results_fields:
                        if any(field.endswith(fv) or field == fv for fv in flag_variants):
                            # Use the documented test name instead
                            return f"{doc_test_name} - {test.get('name', issue_type)}"
                            
                # If no specific test found, use the general test name from documentation
                return f"{doc_test_name}: {issue_type}"
        
        # Fallback to formatted name if no documentation found
        return f"{formatted_test_name}: {issue_type}"

    def format_json_as_table(self, data, indent=0):
        """Convert JSON data to a readable table format"""
        if not isinstance(data, (dict, list)):
            return str(data)

        lines = []
        indent_str = "  " * indent

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    lines.append(f"{indent_str}{key}:")
                    lines.append(self.format_json_as_table(value, indent + 1))
                else:
                    lines.append(f"{indent_str}{key}: {value}")
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, (dict, list)):
                    lines.append(self.format_json_as_table(item, indent))
                else:
                    lines.append(f"{indent_str}- {item}")

        return "\n".join(lines)

    def calculate_summary(self, test_run_ids=None):
        """Calculate summary statistics across all test runs"""
        # If no test_run_ids provided, get all of them
        if test_run_ids is None:
            test_run_ids = self.db.get_all_test_run_ids()
        
        # Build a query that includes all the specified test runs
        query = {}
        if test_run_ids:
            if isinstance(test_run_ids, list):
                query['test_run_id'] = {'$in': test_run_ids}
            else:
                query['test_run_id'] = test_run_ids
        
        results = list(self.db.page_results.find(query))
        
        summary = {
            'total_pages': len(results),
            'pages_with_issues': 0,
            'total_issues': 0,
            'issues_by_type': {},
            'completion_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        for result in results:
            if 'results' in result and 'accessibility' in result['results']:
                tests = result['results']['accessibility'].get('tests', {})
                
                for test_name, test_data in tests.items():
                    if isinstance(test_data, dict):
                        test_obj = test_data.get(test_name, {})
                        page_flags = test_obj.get('pageFlags', {})
                        
                        if page_flags:
                            for flag_name, flag_value in page_flags.items():
                                if isinstance(flag_value, bool) and flag_value and flag_name.startswith('has'):
                                    full_issue_type = self.format_issue_name(test_name, flag_name)
                                    
                                    details = page_flags.get('details', {})
                                    count = 1
                                    
                                    for detail_key, detail_value in details.items():
                                        if isinstance(detail_value, (list, int)):
                                            if isinstance(detail_value, list):
                                                count = len(detail_value)
                                            else:
                                                count = detail_value
                                            
                                            if count > 0:
                                                summary['total_issues'] += count
                                                summary['issues_by_type'][full_issue_type] = \
                                                    summary['issues_by_type'].get(full_issue_type, 0) + count

        summary['pages_with_issues'] = len([
            r for r in results 
            if 'results' in r and 'accessibility' in r['results'] and r['results']['accessibility'].get('tests', {})
        ])

        return summary

    def format_detailed_results(self, results):
        """Format detailed results for Excel with improved JSON formatting"""
        # Each page will have multiple columns based on breakpoints
        
        # First, collect all URLs and all breakpoints
        urls = []
        all_breakpoints = set()
        
        for result in results:
            url = result.get('url', 'Unknown URL')
            urls.append(url)
            accessibility = result.get('results', {}).get('accessibility', {})
            
            if accessibility and 'responsive_testing' in accessibility:
                resp_testing = accessibility['responsive_testing']
                if 'breakpoints' in resp_testing:
                    for bp in resp_testing['breakpoints']:
                        all_breakpoints.add(str(bp))
        
        # Sort breakpoints numerically
        sorted_breakpoints = sorted(all_breakpoints, key=lambda x: int(x) if x.isdigit() else 0)
        
        # Create breakpoint-specific columns for each URL
        url_bp_columns = {}
        for url in urls:
            # Base URL column without breakpoint
            url_bp_columns[url] = url
            
            # Create URL-breakpoint columns for each breakpoint
            for bp in sorted_breakpoints:
                url_bp_columns[f"{url}:{bp}"] = f"{url} @ {bp}px"
        
        # Initialize data dictionary with test names as keys
        formatted_data = {}
        
        # Process regular (non-responsive) test results
        for result in results:
            url = result.get('url', 'Unknown URL')
            accessibility = result.get('results', {}).get('accessibility', {})
            
            if accessibility and 'tests' in accessibility:
                for test_name, test_results in accessibility['tests'].items():
                    # Create a key for this test
                    if test_name not in formatted_data:
                        formatted_data[test_name] = {}
                    
                    # Process dictionary test results
                    if isinstance(test_results, dict):
                        # Flatten the structure for easier Excel formatting
                        flat_data = self._flatten_dict(test_results, test_name)
                        for key, value in flat_data.items():
                            if key not in formatted_data:
                                formatted_data[key] = {}
                            formatted_data[key][url_bp_columns[url]] = value
                    else:
                        # Simple test result
                        formatted_data[test_name][url_bp_columns[url]] = str(test_results)
        
        # Process responsive testing results with breakpoint-specific columns
        for result in results:
            url = result.get('url', 'Unknown URL')
            accessibility = result.get('results', {}).get('accessibility', {})
            
            if accessibility and 'responsive_testing' in accessibility:
                resp_testing = accessibility['responsive_testing']
                
                # Add basic responsive testing summary data
                if 'responsive_testing.summary' not in formatted_data:
                    formatted_data['responsive_testing.summary'] = {}
                
                # Add breakpoints tested
                if 'breakpoints' in resp_testing:
                    formatted_data['responsive_testing.breakpoints_tested'] = formatted_data.get('responsive_testing.breakpoints_tested', {})
                    formatted_data['responsive_testing.breakpoints_tested'][url_bp_columns[url]] = str(resp_testing['breakpoints'])
                
                # Add consolidated summary information
                if 'consolidated' in resp_testing and 'summary' in resp_testing['consolidated']:
                    summary = resp_testing['consolidated']['summary']
                    for key, value in summary.items():
                        summary_key = f"responsive_testing.summary.{key}"
                        if summary_key not in formatted_data:
                            formatted_data[summary_key] = {}
                        formatted_data[summary_key][url_bp_columns[url]] = str(value)
                
                # Process breakpoint-specific results
                for bp_str, bp_results in resp_testing.get('breakpoint_results', {}).items():
                    # Create the breakpoint-specific column key
                    bp_url = f"{url}:{bp_str}"
                    if bp_url not in url_bp_columns:
                        # Skip if this breakpoint wasn't discovered earlier
                        continue
                        
                    if 'tests' in bp_results and 'responsive' in bp_results['tests']:
                        resp_tests = bp_results['tests']['responsive']
                        if 'tests' in resp_tests:
                            # Process each test type at this breakpoint
                            for test_name, test_data in resp_tests['tests'].items():
                                # Create keys for this responsive test
                                resp_key = f"responsive.{test_name}"
                                if resp_key not in formatted_data:
                                    formatted_data[resp_key] = {}
                                
                                # Add issue information
                                issues = test_data.get('issues', [])
                                issue_count = len(issues)
                                
                                if issue_count > 0:
                                    # Summarize the issues
                                    details = []
                                    for issue in issues:
                                        element = f"{issue.get('element', '')} {issue.get('id', '')}".strip()
                                        detail = issue.get('details', '')
                                        severity = issue.get('severity', '')
                                        details.append(f"{element}: {detail} ({severity})")
                                    
                                    # Add to the column for this URL + breakpoint
                                    issue_summary = f"{issue_count} issue(s): {'; '.join(details[:2])}"
                                    if len(details) > 2:
                                        issue_summary += f" and {len(details) - 2} more"
                                    formatted_data[resp_key][url_bp_columns[bp_url]] = issue_summary
                                else:
                                    # No issues found
                                    formatted_data[resp_key][url_bp_columns[bp_url]] = "No issues"
        
        # Convert to DataFrame
        df = pd.DataFrame.from_dict(formatted_data, orient='index')
        
        # Sort columns by URL and breakpoint
        # URLs without breakpoints first, then same URLs with breakpoints in ascending order
        def column_sort_key(col):
            if ' @ ' in col:
                url, bp = col.split(' @ ')
                bp = bp.replace('px', '')
                return (url, int(bp) if bp.isdigit() else 0)
            return (col, 0)  # URLs without breakpoints sort first
        
        sorted_columns = sorted(df.columns, key=column_sort_key)
        if sorted_columns:
            df = df[sorted_columns]
        
        df = df.sort_index()
        return df
        
    def _flatten_dict(self, d, prefix=''):
        """Helper to flatten a nested dictionary with full path keys"""
        items = {}
        for key, value in d.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                items.update(self._flatten_dict(value, full_key))
            elif isinstance(value, list):
                items[full_key] = json.dumps(value)
            else:
                items[full_key] = str(value)
        return items

    def generate_excel_report(self, test_run_ids=None, output_file='accessibility_report.xlsx', db_name=None):
        """Generate Excel report with multiple sheets using all test runs"""
        # If no test_run_ids provided, get all of them
        if test_run_ids is None:
            test_run_ids = self.db.get_all_test_run_ids()
            
        # Prepend database name to output file if provided
        if db_name:
            # Extract the file extension and base name
            if '.' in output_file:
                base, ext = output_file.rsplit('.', 1)
                output_file = f"{db_name}_{base}.{ext}"
            else:
                output_file = f"{db_name}_{output_file}"
        
        # Build a query that includes all the specified test runs
        query = {}
        if test_run_ids:
            if isinstance(test_run_ids, list):
                query['test_run_id'] = {'$in': test_run_ids}
            else:
                query['test_run_id'] = test_run_ids
        
        results = list(self.db.page_results.find(query))
        
        # First, explicitly fetch test documentation from test_runs collection
        # This ensures we get documentation for all tests, even those not in the current results
        print("\n=== DETAILED DOCUMENTATION COLLECTION DEBUG ===")
        print("Explicitly fetching documentation from test_runs collection...")
        try:
            # Get all test runs to look for documentation
            all_test_runs = list(self.db.test_runs.find({}))
            print(f"Found {len(all_test_runs)} test runs in the database")
            
            # Look for specific tests we want to document
            target_tests = ['animations', 'colors', 'forms']
            print(f"Specifically looking for documentation for: {', '.join(target_tests)}")
            
            for run in all_test_runs:
                print(f"\nInspecting test run: {run.get('_id')} ({run.get('name', 'unnamed run')})")
                
                # Debug the test run
                if 'documentation' in run:
                    doc_count = len(run['documentation'])
                    print(f"  Found documentation with {doc_count} test types")
                    print(f"  Documentation keys: {list(run['documentation'].keys())}")
                    
                    # Check for our specific targets
                    for target in target_tests:
                        if target in run['documentation']:
                            print(f"  ✓ Found '{target}' documentation in this test run")
                        elif any(k.startswith(target) or k.endswith(target) for k in run['documentation'].keys()):
                            matching_keys = [k for k in run['documentation'].keys() if k.startswith(target) or k.endswith(target)]
                            print(f"  ⚠ Found similar keys to '{target}': {matching_keys}")
                        else:
                            print(f"  ✗ Did not find '{target}' documentation in this test run")
                    
                    # Process all documentation
                    for test_name, doc in run['documentation'].items():
                        # Store multiple variants of the name to improve matching
                        test_keys = [
                            test_name.lower(),
                            test_name.replace('-', '_').lower(),
                            test_name.replace('_', '-').lower()
                        ]
                        
                        # Use the best key (original first)
                        for test_key in test_keys:
                            if test_key not in self.test_documentation:
                                print(f"  Adding documentation for {test_name} with key {test_key}")
                                self.test_documentation[test_key] = doc
                                # Also add variant keys for improved matching
                                if test_key != test_name.lower():
                                    self.test_documentation[test_name.lower()] = doc
                                break
                            
                # Check for inline documentation in test results
                if 'tests' in run:
                    print(f"  This run has a 'tests' field with {len(run.get('tests', {}))} items")
                    tests = run.get('tests', {})
                    for test_name, test_data in tests.items():
                        if isinstance(test_data, dict) and 'documentation' in test_data:
                            print(f"  Found inline documentation for {test_name} in tests field")
                            self.test_documentation[test_name.lower()] = test_data['documentation']
                            
        except Exception as e:
            print(f"Error fetching documentation from test_runs: {e}")
            import traceback
            traceback.print_exc()
        
        # Then, collect test documentation from results
        self.collect_test_documentation(results)
        
        # For debugging: print all test documentation found
        print("\n=== FINAL DOCUMENTATION SUMMARY ===")
        print(f"Total test types with documentation: {len(self.test_documentation)}")
        
        # Check specifically for our target tests
        target_tests = ['animations', 'colors', 'forms']
        for target in target_tests:
            found = False
            for key in self.test_documentation.keys():
                if key == target or key.startswith(target) or key.endswith(target):
                    found = True
                    print(f"✓ '{target}' documentation found as key '{key}'")
                    break
            if not found:
                print(f"✗ '{target}' documentation NOT FOUND in any key")
        
        print("\nAll documented test types:")
        for test_name in sorted(self.test_documentation.keys()):
            test_obj = self.test_documentation[test_name]
            display_name = test_obj.get('testName', test_name)
            tests_count = len(test_obj.get('tests', []))
            print(f"  - {test_name} → {display_name} ({tests_count} subtests)")
        
        summary = self.calculate_summary(test_run_ids)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_df = pd.DataFrame({
                'Metric': [
                    'Total Pages',
                    'Pages with Issues',
                    'Total Issues',
                    'Completion Time'
                ],
                'Value': [
                    summary['total_pages'],
                    summary['pages_with_issues'],
                    summary['total_issues'],
                    summary['completion_time']
                ]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Issues by type
            if summary['issues_by_type']:
                issues_df = pd.DataFrame(
                    summary['issues_by_type'].items(),
                    columns=['Issue Type', 'Count']
                )
                issues_df.to_excel(writer, sheet_name='Issues by Type', index=False)
            
            # Issues by URL
            issues_by_url = {}
            for result in results:
                url = result.get('url', 'Unknown URL')
                if 'results' in result and 'accessibility' in result['results']:
                    tests = result['results']['accessibility']['tests']
                    url_issues = []
                    
                    for test_name, test_data in tests.items():
                        if isinstance(test_data, dict):
                            test_obj = test_data.get(test_name, {})
                            page_flags = test_obj.get('pageFlags', {})
                            
                            for flag_name, flag_value in page_flags.items():
                                if isinstance(flag_value, bool) and flag_value and flag_name.startswith('has'):
                                    full_issue_type = self.format_issue_name(test_name, flag_name)
                                    
                                    details = page_flags.get('details', {})
                                    count = 1
                                    for detail_key, detail_value in details.items():
                                        if isinstance(detail_value, (list, int)):
                                            if isinstance(detail_value, list):
                                                count = len(detail_value)
                                            else:
                                                count = detail_value
                                            if count > 0:
                                                url_issues.append({
                                                    'Issue Type': full_issue_type,
                                                    'Count': count,
                                                    'Details': f"Found {count} issue(s)"
                                                })
                    
                    if url_issues:
                        issues_by_url[url] = url_issues
            
            if issues_by_url:
                url_issues_data = []
                for url, issues in issues_by_url.items():
                    for issue in issues:
                        url_issues_data.append({
                            'URL': url,
                            'Issue Type': issue['Issue Type'],
                            'Count': issue['Count'],
                            'Details': issue['Details']
                        })
                
                url_issues_df = pd.DataFrame(url_issues_data)
                url_issues_df.to_excel(writer, sheet_name='Issues by URL', index=False)
            
            # Issues by Site
            issues_by_site = {}
            for result in results:
                full_url = result.get('url', 'Unknown URL')
                try:
                    parsed_url = urlparse(full_url)
                    site_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
                except:
                    site_url = 'Unknown Site'

                if 'results' in result and 'accessibility' in result['results']:
                    tests = result['results']['accessibility']['tests']
                    
                    if site_url not in issues_by_site:
                        issues_by_site[site_url] = {}
                    
                    for test_name, test_data in tests.items():
                        if isinstance(test_data, dict):
                            test_obj = test_data.get(test_name, {})
                            page_flags = test_obj.get('pageFlags', {})
                            
                            for flag_name, flag_value in page_flags.items():
                                if isinstance(flag_value, bool) and flag_value and flag_name.startswith('has'):
                                    full_issue_type = self.format_issue_name(test_name, flag_name)
                                    
                                    details = page_flags.get('details', {})
                                    count = 1
                                    for detail_key, detail_value in details.items():
                                        if isinstance(detail_value, (list, int)):
                                            if isinstance(detail_value, list):
                                                count = len(detail_value)
                                            else:
                                                count = detail_value
                                            if count > 0:
                                                if full_issue_type not in issues_by_site[site_url]:
                                                    issues_by_site[site_url][full_issue_type] = {
                                                        'Count': 0,
                                                        'Pages Affected': 0
                                                    }
                                                issues_by_site[site_url][full_issue_type]['Count'] += count
                                                issues_by_site[site_url][full_issue_type]['Pages Affected'] += 1

            if issues_by_site:
                site_issues_data = []
                for site_url, issues in issues_by_site.items():
                    for issue_type, data in issues.items():
                        site_issues_data.append({
                            'Site': site_url,
                            'Issue Type': issue_type,
                            'Total Count': data['Count'],
                            'Pages Affected': data['Pages Affected']
                        })
                
                site_issues_df = pd.DataFrame(site_issues_data)
                site_issues_df.to_excel(writer, sheet_name='Issues by Site', index=False)
            
            # Documentation sheet
            if self.test_documentation:
                print("\nPreparing documentation sheet...")
                docs_data = []
                for test_name, doc in self.test_documentation.items():
                    # Get a clean test name for display
                    display_test_name = doc.get('testName', test_name.replace('_', ' ').title())
                    
                    print(f"Adding documentation for {display_test_name}")
                    
                    # Add test general documentation
                    docs_data.append({
                        'Test Name': display_test_name,
                        'Type': 'Test',
                        'Description': doc.get('description', ''),
                        'Version': doc.get('version', ''),
                        'Date': doc.get('date', ''),
                        'WCAG Criteria': '',
                        'Impact': '',
                        'How to Fix': ''
                    })
                    
                    # Add individual checks documentation
                    for test in doc.get('tests', []):
                        subtest_name = test.get('name', '')
                        if subtest_name:  # Only add subtests that have names
                            full_name = f"{display_test_name} - {subtest_name}"
                            print(f"  Adding subtest: {subtest_name}")
                            
                            docs_data.append({
                                'Test Name': full_name,
                                'Type': 'Check',
                                'Description': test.get('description', ''),
                                'Version': doc.get('version', ''),
                                'Date': doc.get('date', ''),
                                'WCAG Criteria': ', '.join(test.get('wcagCriteria', [])),
                                'Impact': test.get('impact', ''),
                                'How to Fix': test.get('howToFix', '')
                            })
                
                if docs_data:
                    # Sort by Test Name and make Type a secondary sort key (Test first, then Check)
                    # This ensures each test is followed by its checks
                    print(f"Creating documentation sheet with {len(docs_data)} entries")
                    docs_df = pd.DataFrame(docs_data)
                    
                    # Create a custom sorter for Type column
                    type_order = {'Test': 0, 'Check': 1}
                    docs_df['type_order'] = docs_df['Type'].map(type_order)
                    
                    # Sort first by Test Name, then by type_order
                    docs_df = docs_df.sort_values(['Test Name', 'type_order'])
                    
                    # Remove the temporary sorting column
                    docs_df = docs_df.drop(columns=['type_order'])
                    
                    # Write to Excel
                    docs_df.to_excel(writer, sheet_name='Test Documentation', index=False)
                    
                    print("Documentation sheet created successfully")
                else:
                    print("No documentation data found to include in report")
            
            # Detailed results
            detailed_df = self.format_detailed_results(results)
            detailed_df.to_excel(writer, sheet_name='Detailed Results')
            
            # Create a new, more readable responsive testing sheet
            responsive_matrix_data = []
            test_types = set()  # Track all test types we find
            
            # First pass - collect all test types
            for result in results:
                accessibility = result.get('results', {}).get('accessibility', {})
                if accessibility and 'responsive_testing' in accessibility:
                    resp_testing = accessibility['responsive_testing']
                    
                    for bp, bp_results in resp_testing.get('breakpoint_results', {}).items():
                        if 'tests' in bp_results and 'responsive' in bp_results['tests'] and 'tests' in bp_results['tests']['responsive']:
                            for test_name in bp_results['tests']['responsive']['tests'].keys():
                                test_types.add(test_name)
            
            # Second pass - create rows for URL + breakpoint + test type
            for result in results:
                url = result.get('url', 'Unknown URL')
                domain_parts = url.replace('https://', '').replace('http://', '').split('/')
                domain = domain_parts[0]
                page = '/'.join(domain_parts[1:]) if len(domain_parts) > 1 else ''
                
                accessibility = result.get('results', {}).get('accessibility', {})
                if accessibility and 'responsive_testing' in accessibility:
                    resp_testing = accessibility['responsive_testing']
                    
                    # Get all breakpoints
                    breakpoints = resp_testing.get('breakpoints', [])
                    
                    # Process each breakpoint
                    for bp in breakpoints:
                        bp_str = str(bp)
                        
                        # Look for results for this breakpoint
                        if bp_str in resp_testing.get('breakpoint_results', {}):
                            bp_results = resp_testing['breakpoint_results'][bp_str]
                            
                            if 'tests' in bp_results and 'responsive' in bp_results['tests'] and 'tests' in bp_results['tests']['responsive']:
                                resp_tests = bp_results['tests']['responsive']['tests']
                                
                                # Process each test type
                                for test_name in sorted(test_types):
                                    # Create a new row for each test at this breakpoint
                                    row_data = {
                                        'Domain': domain,
                                        'Page': page,
                                        'Breakpoint': bp,
                                        'Test': test_name
                                    }
                                    
                                    # Add status and details for this test
                                    if test_name in resp_tests:
                                        test_data = resp_tests[test_name]
                                        issues = test_data.get('issues', [])
                                        issue_count = len(issues)
                                        
                                        if issue_count > 0:
                                            row_data['Status'] = 'Issues Found'
                                            row_data['Issue Count'] = issue_count
                                            
                                            # Extract issue details
                                            issue_details = []
                                            for issue in issues:
                                                element = f"{issue.get('element', '')} {issue.get('id', '')}".strip()
                                                detail = issue.get('details', '')
                                                severity = issue.get('severity', '')
                                                issue_details.append(f"{element}: {detail} ({severity})")
                                            
                                            row_data['Issue Details'] = '\n'.join(issue_details[:3])
                                            if len(issue_details) > 3:
                                                row_data['Issue Details'] += f"\n...and {len(issue_details) - 3} more"
                                        else:
                                            row_data['Status'] = 'Pass'
                                            row_data['Issue Count'] = 0
                                            row_data['Issue Details'] = 'No issues detected'
                                    else:
                                        # Test not run at this breakpoint
                                        row_data['Status'] = 'Not Tested'
                                        row_data['Issue Count'] = '-'
                                        row_data['Issue Details'] = '-'
                                    
                                    # Add the row
                                    responsive_matrix_data.append(row_data)
            
            # Create the responsive matrix sheet
            if responsive_matrix_data:
                # Sort data by domain, page, breakpoint, test
                responsive_matrix_data.sort(key=lambda x: (x['Domain'], x['Page'], x['Breakpoint'], x['Test']))
                resp_matrix_df = pd.DataFrame(responsive_matrix_data)
                resp_matrix_df.to_excel(writer, sheet_name='Responsive Testing Matrix', index=False)
                
            # Create a breakpoint summary sheet
            breakpoint_summary_data = []
            
            # Track all distinct breakpoints and issue counts for visualization
            all_breakpoints = set()
            issue_by_breakpoint = defaultdict(int)
            issue_by_type = defaultdict(int)
            domains_with_issues = set()
            
            for result in results:
                url = result.get('url', 'Unknown URL')
                domain_parts = url.replace('https://', '').replace('http://', '').split('/')
                domain = domain_parts[0]
                
                accessibility = result.get('results', {}).get('accessibility', {})
                
                if accessibility and 'responsive_testing' in accessibility:
                    resp_testing = accessibility['responsive_testing']
                    
                    # Get all breakpoints tested
                    breakpoints = resp_testing.get('breakpoints', [])
                    for bp in breakpoints:
                        all_breakpoints.add(int(bp))
                    
                    # Get consolidated results for test summary
                    if 'consolidated' in resp_testing:
                        consolidated = resp_testing['consolidated']
                        
                        # Record issue types for chart
                        if 'summary' in consolidated:
                            summary = consolidated['summary']
                            issue_by_type['Overflow'] += summary.get('overflowIssues', 0)
                            issue_by_type['Touch Target'] += summary.get('touchTargetIssues', 0)
                            issue_by_type['Font Scaling'] += summary.get('fontScalingIssues', 0)
                            issue_by_type['Fixed Position'] += summary.get('fixedPositionIssues', 0)
                            issue_by_type['Content Stacking'] += summary.get('contentStackingIssues', 0)
                        
                        # Add summary row for this URL
                        breakpoint_summary_data.append({
                            'URL': url,
                            'Total Breakpoints Tested': len(breakpoints),
                            'Breakpoints with Issues': consolidated.get('summary', {}).get('affectedBreakpoints', 0),
                            'Total Issues': consolidated.get('summary', {}).get('totalIssues', 0),
                            'Overflow Issues': consolidated.get('summary', {}).get('overflowIssues', 0),
                            'Touch Target Issues': consolidated.get('summary', {}).get('touchTargetIssues', 0),
                            'Font Scaling Issues': consolidated.get('summary', {}).get('fontScalingIssues', 0),
                            'Fixed Position Issues': consolidated.get('summary', {}).get('fixedPositionIssues', 0),
                            'Content Stacking Issues': consolidated.get('summary', {}).get('contentStackingIssues', 0),
                            'Breakpoints Tested': ', '.join(map(str, sorted(breakpoints)))
                        })
                        
                        # Track if domain has issues
                        if consolidated.get('summary', {}).get('totalIssues', 0) > 0:
                            domains_with_issues.add(domain)
                        
                        # Add breakpoint-specific rows
                        for bp, bp_results in resp_testing.get('breakpoint_results', {}).items():
                            if 'tests' in bp_results and 'responsive' in bp_results['tests']:
                                resp_data = bp_results['tests']['responsive']
                                
                                # Count total issues for this breakpoint
                                total_issues = 0
                                issue_details = []
                                
                                # Process each test type
                                for test_name, test_data in resp_data.get('tests', {}).items():
                                    if 'issues' in test_data:
                                        issues_count = len(test_data['issues'])
                                        if issues_count > 0:
                                            total_issues += issues_count
                                            issue_details.append(f"{test_name}: {issues_count}")
                                
                                # Record issues by breakpoint for charts
                                if total_issues > 0:
                                    issue_by_breakpoint[int(bp)] += total_issues
                                    
                                    breakpoint_summary_data.append({
                                        'URL': f"  -- {url} @ {bp}px",
                                        'Total Breakpoints Tested': '',
                                        'Breakpoints with Issues': '',
                                        'Total Issues': total_issues,
                                        'Overflow Issues': '',
                                        'Touch Target Issues': '',
                                        'Font Scaling Issues': '',
                                        'Fixed Position Issues': '',
                                        'Content Stacking Issues': '',
                                        'Breakpoints Tested': ', '.join(issue_details)
                                    })
            
            # Only create the breakpoint summary sheet if we have data
            if breakpoint_summary_data:
                bp_df = pd.DataFrame(breakpoint_summary_data)
                bp_df.to_excel(writer, sheet_name='Responsive Breakpoint Summary', index=False)
            
            # Create a visualization sheet if we have responsive data
            if all_breakpoints and (issue_by_breakpoint or issue_by_type):
                # Create a new sheet for visualizations
                workbook = writer.book
                vis_sheet = workbook.create_sheet('Responsive Visualizations')
                
                # Create the data tables for the charts
                vis_sheet['A1'] = 'Issues by Breakpoint'
                vis_sheet['A2'] = 'Breakpoint (px)'
                vis_sheet['B2'] = 'Issue Count'
                
                # Sort breakpoints for chart
                sorted_breakpoints = sorted(all_breakpoints)
                
                # Add breakpoint data
                row = 3
                for bp in sorted_breakpoints:
                    vis_sheet[f'A{row}'] = bp
                    vis_sheet[f'B{row}'] = issue_by_breakpoint.get(bp, 0)
                    row += 1
                
                # Create Breakpoint chart using openpyxl
                chart = openpyxl.chart.BarChart()
                chart.title = "Responsive Accessibility Issues by Breakpoint"
                chart.style = 10  # Choose a style (1-48)
                chart.x_axis.title = "Breakpoint (px)"
                chart.y_axis.title = "Number of Issues"
                
                # Add the data
                bp_data = openpyxl.chart.Reference(
                    vis_sheet, 
                    min_col=2, 
                    min_row=2, 
                    max_row=2+len(sorted_breakpoints)
                )
                bp_labels = openpyxl.chart.Reference(
                    vis_sheet, 
                    min_col=1, 
                    min_row=3, 
                    max_row=2+len(sorted_breakpoints)
                )
                chart.add_data(bp_data, titles_from_data=True)
                chart.set_categories(bp_labels)
                
                # Make the chart larger
                chart.width = 30
                chart.height = 15
                
                # Add the chart to the sheet
                vis_sheet.add_chart(chart, "A15")
                
                # Add issue type data table
                vis_sheet['D1'] = 'Issues by Type'
                vis_sheet['D2'] = 'Issue Type'
                vis_sheet['E2'] = 'Issue Count'
                
                # Add the issue type data
                row = 3
                for issue_type, count in issue_by_type.items():
                    vis_sheet[f'D{row}'] = issue_type
                    vis_sheet[f'E{row}'] = count
                    row += 1
                
                # Create issue type chart
                pie_chart = openpyxl.chart.PieChart()
                pie_chart.title = "Distribution of Responsive Issues by Type"
                pie_chart.style = 10
                
                # Add the data
                type_data = openpyxl.chart.Reference(
                    vis_sheet, 
                    min_col=5, 
                    min_row=2, 
                    max_row=2+len(issue_by_type)
                )
                type_labels = openpyxl.chart.Reference(
                    vis_sheet, 
                    min_col=4, 
                    min_row=3, 
                    max_row=2+len(issue_by_type)
                )
                pie_chart.add_data(type_data, titles_from_data=True)
                pie_chart.set_categories(type_labels)
                
                # Show data labels
                slice_series = pie_chart.series[0]
                slice_series.data_labels = openpyxl.chart.label.DataLabelList()
                slice_series.data_labels.showVal = False
                slice_series.data_labels.showPercent = True
                slice_series.data_labels.showCatName = True
                
                # Make the chart larger
                pie_chart.width = 20
                pie_chart.height = 15
                
                # Add the chart to the sheet
                vis_sheet.add_chart(pie_chart, "D15")
                
                # Create a heatmap-like table showing breakpoints with issues for all domains
                if domains_with_issues:
                    vis_sheet['G1'] = 'Breakpoint Issues Heatmap by Domain'
                    vis_sheet['G1'].font = openpyxl.styles.Font(size=14, bold=True)
                    
                    # Add breakpoint headers
                    col = 8  # Start at column H (col index 8)
                    for bp in sorted_breakpoints:
                        cell = vis_sheet.cell(row=2, column=col)
                        cell.value = f"{bp}px"
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color="E6E6E6",
                            end_color="E6E6E6",
                            fill_type="solid"
                        )
                        col += 1
                    
                    # Track domain-breakpoint issue counts for heatmap
                    domain_bp_issues = {}
                    
                    # Compile issue counts for each domain at each breakpoint
                    for result in results:
                        url = result.get('url', 'Unknown URL')
                        domain_parts = url.replace('https://', '').replace('http://', '').split('/')
                        domain = domain_parts[0]
                        
                        accessibility = result.get('results', {}).get('accessibility', {})
                        
                        if accessibility and 'responsive_testing' in accessibility:
                            resp_testing = accessibility['responsive_testing']
                            
                            # Initialize domain in tracking dict if needed
                            if domain not in domain_bp_issues:
                                domain_bp_issues[domain] = defaultdict(int)
                            
                            # Process each breakpoint result
                            for bp, bp_results in resp_testing.get('breakpoint_results', {}).items():
                                if 'tests' in bp_results and 'responsive' in bp_results['tests']:
                                    resp_data = bp_results['tests']['responsive']
                                    
                                    # Count total issues for this breakpoint
                                    total_issues = 0
                                    
                                    # Process each test type
                                    for test_name, test_data in resp_data.get('tests', {}).items():
                                        if 'issues' in test_data:
                                            total_issues += len(test_data['issues'])
                                    
                                    # Add issues to domain-breakpoint tracking
                                    if total_issues > 0:
                                        domain_bp_issues[domain][int(bp)] += total_issues
                    
                    # Add domain rows with heatmap coloring
                    row = 3
                    for domain in sorted(domains_with_issues):
                        # Add domain cell
                        domain_cell = vis_sheet.cell(row=row, column=7)  # Column G
                        domain_cell.value = domain
                        domain_cell.font = openpyxl.styles.Font(bold=True)
                        
                        # Add cells for each breakpoint with heatmap coloring
                        col = 8  # Start at column H
                        for bp in sorted_breakpoints:
                            cell = vis_sheet.cell(row=row, column=col)
                            
                            # Get issue count for this domain-breakpoint combination
                            issue_count = domain_bp_issues.get(domain, {}).get(bp, 0)
                            
                            # Set the value
                            cell.value = issue_count if issue_count > 0 else ""
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            
                            # Apply heatmap coloring based on issue count
                            if issue_count > 10:  # High severity
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FF9999",  # Darker red
                                    end_color="FF9999",
                                    fill_type="solid"
                                )
                                cell.font = openpyxl.styles.Font(bold=True)
                            elif issue_count > 5:  # Medium-high severity
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFCCCC",  # Light red
                                    end_color="FFCCCC",
                                    fill_type="solid"
                                )
                            elif issue_count > 2:  # Medium severity
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFE0B2",  # Light orange
                                    end_color="FFE0B2",
                                    fill_type="solid"
                                )
                            elif issue_count > 0:  # Low severity
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFF9C4",  # Light yellow
                                    end_color="FFF9C4",
                                    fill_type="solid"
                                )
                            
                            col += 1
                        
                        row += 1
                    
                    # Add a legend for the heatmap
                    legend_row = row + 2
                    vis_sheet.cell(row=legend_row, column=7).value = "Heatmap Legend:"
                    vis_sheet.cell(row=legend_row, column=7).font = openpyxl.styles.Font(bold=True)
                    
                    # High severity
                    vis_sheet.cell(row=legend_row, column=8).fill = openpyxl.styles.PatternFill(
                        start_color="FF9999", end_color="FF9999", fill_type="solid")
                    vis_sheet.cell(row=legend_row, column=9).value = "> 10 issues (High)"
                    
                    # Medium-high severity
                    vis_sheet.cell(row=legend_row+1, column=8).fill = openpyxl.styles.PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    vis_sheet.cell(row=legend_row+1, column=9).value = "6-10 issues (Medium-High)"
                    
                    # Medium severity
                    vis_sheet.cell(row=legend_row+2, column=8).fill = openpyxl.styles.PatternFill(
                        start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                    vis_sheet.cell(row=legend_row+2, column=9).value = "3-5 issues (Medium)"
                    
                    # Low severity
                    vis_sheet.cell(row=legend_row+3, column=8).fill = openpyxl.styles.PatternFill(
                        start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    vis_sheet.cell(row=legend_row+3, column=9).value = "1-2 issues (Low)"
                    
                    # Add trend analysis - issues per test type across breakpoints
                    trend_row = legend_row + 6
                    vis_sheet.cell(row=trend_row, column=7).value = "Issue Trends by Test Type"
                    vis_sheet.cell(row=trend_row, column=7).font = openpyxl.styles.Font(size=14, bold=True)
                    
                    # Collect issue data by test type across breakpoints
                    test_types = ['overflow', 'touchTargets', 'fontScaling', 'fixedPosition', 'contentStacking']
                    test_type_labels = {
                        'overflow': 'Content Overflow', 
                        'touchTargets': 'Touch Targets',
                        'fontScaling': 'Font Scaling',
                        'fixedPosition': 'Fixed Position',
                        'contentStacking': 'Content Stacking'
                    }
                    
                    # Create the trend table headers with breakpoints
                    trend_row += 1
                    vis_sheet.cell(row=trend_row, column=7).value = "Test Type"
                    col = 8
                    for bp in sorted_breakpoints:
                        vis_sheet.cell(row=trend_row, column=col).value = f"{bp}px"
                        vis_sheet.cell(row=trend_row, column=col).font = openpyxl.styles.Font(bold=True)
                        vis_sheet.cell(row=trend_row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                        vis_sheet.cell(row=trend_row, column=col).fill = openpyxl.styles.PatternFill(
                            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                        col += 1
                    
                    # Collect data by test type across breakpoints
                    test_type_data = {}
                    for test_type in test_types:
                        test_type_data[test_type] = defaultdict(int)
                    
                    # Process all results to collect data
                    for result in results:
                        accessibility = result.get('results', {}).get('accessibility', {})
                        if accessibility and 'responsive_testing' in accessibility:
                            resp_testing = accessibility['responsive_testing']
                            
                            for bp, bp_results in resp_testing.get('breakpoint_results', {}).items():
                                if 'tests' in bp_results and 'responsive' in bp_results['tests']:
                                    resp_data = bp_results['tests']['responsive']
                                    
                                    for test_name, test_data in resp_data.get('tests', {}).items():
                                        if test_name in test_types and 'issues' in test_data:
                                            test_type_data[test_name][int(bp)] += len(test_data['issues'])
                    
                    # Add each test type row
                    trend_row += 1
                    for test_type in test_types:
                        vis_sheet.cell(row=trend_row, column=7).value = test_type_labels.get(test_type, test_type)
                        vis_sheet.cell(row=trend_row, column=7).font = openpyxl.styles.Font(bold=True)
                        
                        col = 8
                        for bp in sorted_breakpoints:
                            # Get issue count for this test type and breakpoint
                            issue_count = test_type_data[test_type][bp]
                            
                            # Add to the cell
                            cell = vis_sheet.cell(row=trend_row, column=col)
                            cell.value = issue_count if issue_count > 0 else ""
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            
                            # Apply coloring based on count
                            if issue_count > 0:
                                # Use test-specific color scheme
                                if test_type == 'overflow':
                                    base_color = "FFCCCC"  # Red
                                elif test_type == 'touchTargets':
                                    base_color = "FFE0B2"  # Orange
                                elif test_type == 'fontScaling':
                                    base_color = "FFF9C4"  # Yellow
                                elif test_type == 'fixedPosition':
                                    base_color = "C8E6C9"  # Green
                                elif test_type == 'contentStacking':
                                    base_color = "BBDEFB"  # Blue
                                else:
                                    base_color = "E1BEE7"  # Purple
                                
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color=base_color,
                                    end_color=base_color,
                                    fill_type="solid"
                                )
                                
                                # Bold for higher counts
                                if issue_count > 5:
                                    cell.font = openpyxl.styles.Font(bold=True)
                            
                            col += 1
                        
                        trend_row += 1
                    
                    # Create a column chart instead of line chart - simpler to work with
                    trend_row += 2
                    col_chart = openpyxl.chart.BarChart()
                    col_chart.type = "col"
                    col_chart.title = "Responsive Issues by Test Type Across Breakpoints"
                    col_chart.style = 12
                    col_chart.x_axis.title = "Breakpoint (px)"
                    col_chart.y_axis.title = "Number of Issues"
                    col_chart.grouping = "stacked"
                    
                    # First, let's ensure our labels are in the worksheet
                    # Add a column for labels
                    label_col = 7
                    for i, test_type in enumerate(test_types):
                        label_row = trend_row - len(test_types) + i
                        vis_sheet.cell(row=label_row, column=label_col).value = test_type_labels.get(test_type, test_type)
                    
                    # Create data for chart
                    data = openpyxl.chart.Reference(
                        vis_sheet,
                        min_col=8,
                        max_col=8 + len(sorted_breakpoints) - 1,
                        min_row=trend_row - len(test_types),
                        max_row=trend_row - 1
                    )
                    
                    # Create categories for X axis (breakpoints)
                    cats = openpyxl.chart.Reference(
                        vis_sheet,
                        min_col=8,
                        max_col=8 + len(sorted_breakpoints) - 1,
                        min_row=trend_row - len(test_types) - 1,
                        max_row=trend_row - len(test_types) - 1
                    )
                    
                    # Create series titles
                    series_titles = openpyxl.chart.Reference(
                        vis_sheet,
                        min_col=7,
                        max_col=7,
                        min_row=trend_row - len(test_types),
                        max_row=trend_row - 1
                    )
                    
                    # Add the data
                    col_chart.add_data(data, titles_from_data=False)
                    col_chart.set_categories(cats)
                    col_chart.dataLabels = openpyxl.chart.label.DataLabelList()
                    col_chart.dataLabels.showVal = True
                    
                    # Don't try to set the series titles directly - use a separate legend instead
                    # Create a legend with proper positioning
                    col_chart.legend = openpyxl.chart.legend.Legend()
                    col_chart.legend.position = 'r'  # Position legend to the right
                    
                    # Make the chart larger
                    col_chart.width = 30
                    col_chart.height = 15
                    
                    # Add the chart to the sheet
                    vis_sheet.add_chart(col_chart, "A40")
                    
                    # Add note about the trend analysis
                    trend_note_row = trend_row + 20
                    vis_sheet.cell(row=trend_note_row, column=7).value = "Note: The trend analysis shows which test types have the most issues at each breakpoint."
                    vis_sheet.cell(row=trend_note_row, column=7).font = openpyxl.styles.Font(italic=True)
                    
                    # Add mobile vs desktop comparison section
                    mobile_vs_desktop_row = trend_note_row + 3
                    vis_sheet.cell(row=mobile_vs_desktop_row, column=7).value = "Mobile vs Desktop Comparison"
                    vis_sheet.cell(row=mobile_vs_desktop_row, column=7).font = openpyxl.styles.Font(size=14, bold=True)
                    
                    # Create the comparison table headers
                    vis_sheet.cell(row=mobile_vs_desktop_row+1, column=7).value = "Issue Type"
                    vis_sheet.cell(row=mobile_vs_desktop_row+1, column=8).value = "Mobile (<= 768px)"
                    vis_sheet.cell(row=mobile_vs_desktop_row+1, column=9).value = "Desktop (> 768px)"
                    
                    # Format headers
                    for col in range(7, 10):
                        cell = vis_sheet.cell(row=mobile_vs_desktop_row+1, column=col)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    
                    # Calculate mobile vs desktop issues by test type
                    mobile_desktop_data = {}
                    for test_type in test_types:
                        mobile_desktop_data[test_type] = {'mobile': 0, 'desktop': 0}
                        
                        for bp, count in test_type_data[test_type].items():
                            if bp <= 768:
                                mobile_desktop_data[test_type]['mobile'] += count
                            else:
                                mobile_desktop_data[test_type]['desktop'] += count
                    
                    # Add data rows
                    for i, test_type in enumerate(test_types):
                        row = mobile_vs_desktop_row + 2 + i
                        vis_sheet.cell(row=row, column=7).value = test_type_labels.get(test_type, test_type)
                        vis_sheet.cell(row=row, column=8).value = mobile_desktop_data[test_type]['mobile']
                        vis_sheet.cell(row=row, column=9).value = mobile_desktop_data[test_type]['desktop']
                        
                        # Color-code cells based on values
                        for j, col_idx in enumerate([8, 9]):
                            device_type = 'mobile' if j == 0 else 'desktop'
                            count = mobile_desktop_data[test_type][device_type]
                            cell = vis_sheet.cell(row=row, column=col_idx)
                            
                            # Apply coloring based on count
                            if count > 10:
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FF9999", end_color="FF9999", fill_type="solid")
                                cell.font = openpyxl.styles.Font(bold=True)
                            elif count > 5:
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                            elif count > 0:
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    
                    # Create a bar chart comparing mobile vs desktop issues
                    compare_chart = openpyxl.chart.BarChart()
                    compare_chart.type = "col"
                    compare_chart.title = "Mobile vs Desktop Comparison"
                    compare_chart.style = 12
                    compare_chart.x_axis.title = "Issue Type"
                    compare_chart.y_axis.title = "Number of Issues"
                    
                    # Data for the chart
                    compare_data = openpyxl.chart.Reference(
                        vis_sheet,
                        min_col=8,
                        max_col=9,
                        min_row=mobile_vs_desktop_row + 1,
                        max_row=mobile_vs_desktop_row + 1 + len(test_types)
                    )
                    
                    # Categories
                    compare_cats = openpyxl.chart.Reference(
                        vis_sheet,
                        min_col=7,
                        max_col=7,
                        min_row=mobile_vs_desktop_row + 2,
                        max_row=mobile_vs_desktop_row + 1 + len(test_types)
                    )
                    
                    # Add data and configure
                    compare_chart.add_data(compare_data, titles_from_data=True)
                    compare_chart.set_categories(compare_cats)
                    compare_chart.shape = 4  # 4 gives cylinder shape bars
                    
                    # Make the chart larger
                    compare_chart.width = 25
                    compare_chart.height = 15
                    
                    # Add the chart to the sheet
                    vis_sheet.add_chart(compare_chart, "D40")
                    
                    # Add a summary note
                    compare_note_row = mobile_vs_desktop_row + len(test_types) + 3
                    vis_sheet.cell(row=compare_note_row, column=7).value = "This comparison helps identify which issues are more prevalent on mobile vs desktop screens."
                    vis_sheet.cell(row=compare_note_row, column=7).font = openpyxl.styles.Font(italic=True)
            
            # Apply formatting to all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Format header row
                for cell in worksheet[1]:
                    cell.font = openpyxl.styles.Font(size=16)
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal='center'
                    )
                    # Add background color to header cells
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="E6E6E6",
                        end_color="E6E6E6",
                        fill_type="solid"
                    )
                worksheet.row_dimensions[1].height = 40
                
                # Add conditional formatting to responsive breakpoint columns in Detailed Results sheet
                if sheet_name == 'Detailed Results':
                    # First identify responsive test rows
                    responsive_test_rows = []
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                        if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('responsive.'):
                            responsive_test_rows.append(row_idx)
                    
                    # Now add conditional formatting to cells in these rows
                    for row_idx in responsive_test_rows:
                        for cell in worksheet[row_idx]:
                            if isinstance(cell.value, str):
                                # Apply color based on content
                                if "issue" in cell.value.lower() and "no issues" not in cell.value.lower():
                                    # Extract issue count if possible
                                    issue_count = 0
                                    try:
                                        if " issue(s)" in cell.value:
                                            count_text = cell.value.split(" issue(s)")[0]
                                            issue_count = int(count_text)
                                    except:
                                        pass
                                    
                                    # Apply severity-based color
                                    if issue_count > 3 or "high" in cell.value.lower():
                                        # High severity - light red
                                        cell.fill = openpyxl.styles.PatternFill(
                                            start_color="FFCCCC",
                                            end_color="FFCCCC",
                                            fill_type="solid"
                                        )
                                    elif issue_count > 1 or "medium" in cell.value.lower():
                                        # Medium severity - light orange
                                        cell.fill = openpyxl.styles.PatternFill(
                                            start_color="FFE0B2",
                                            end_color="FFE0B2",
                                            fill_type="solid"
                                        )
                                    else:
                                        # Low severity - light yellow
                                        cell.fill = openpyxl.styles.PatternFill(
                                            start_color="FFF9C4",
                                            end_color="FFF9C4",
                                            fill_type="solid"
                                        )
                                elif "no issues" in cell.value.lower():
                                    # No issues - light green
                                    cell.fill = openpyxl.styles.PatternFill(
                                        start_color="E8F5E9",
                                        end_color="E8F5E9",
                                        fill_type="solid"
                                    )
                
                # Apply conditional formatting to the Responsive Testing Matrix sheet
                if sheet_name == 'Responsive Testing Matrix':
                    # Add column headers highlighting
                    for row in worksheet.iter_rows(min_row=2):
                        # Process the status column
                        status_col_idx = None
                        for i, cell in enumerate(row):
                            if i == 0:  # Domain column - add light gray fill
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="F5F5F5",
                                    end_color="F5F5F5",
                                    fill_type="solid"
                                )
                            if cell.value == 'Status':
                                status_col_idx = i
                                break
                            
                            if cell.value == 'Status':
                                status_col_idx = i
                            
                            # Format the Breakpoint column - add subtle fill to identify breakpoint changes
                            if i == 2 and cell.value and cell.column_letter:  # Breakpoint column
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="F5F5F5", 
                                    end_color="F5F5F5",
                                    fill_type="solid"
                                )
                        
                        # Process status cell if found
                        if status_col_idx is not None:
                            status_cell = row[status_col_idx]
                            if status_cell.value == 'Issues Found':
                                status_cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFCCCC",  # Light red
                                    end_color="FFCCCC",
                                    fill_type="solid"
                                )
                                status_cell.font = openpyxl.styles.Font(size=16, bold=True)
                            elif status_cell.value == 'Pass':
                                status_cell.fill = openpyxl.styles.PatternFill(
                                    start_color="E8F5E9",  # Light green
                                    end_color="E8F5E9",
                                    fill_type="solid"
                                )
                
                # Add conditional formatting to the Responsive Breakpoint Summary sheet
                if sheet_name == 'Responsive Breakpoint Summary':
                    for row in worksheet.iter_rows(min_row=2):
                        # Check for the Total Issues column
                        total_issues_col = None
                        for i, cell in enumerate(row):
                            if i == 0 and cell.value and '  -- ' not in str(cell.value):  # Main URL row
                                cell.font = openpyxl.styles.Font(size=16, bold=True)
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="E3F2FD",  # Light blue
                                    end_color="E3F2FD",
                                    fill_type="solid"
                                )
                            elif i == 0 and cell.value and '  -- ' in str(cell.value):  # Breakpoint subrow
                                # Extract the breakpoint from the URL @ {bp}px format
                                try:
                                    bp_text = cell.value.split('@')[1].strip()
                                    bp_px = int(bp_text.replace('px', ''))
                                    
                                    # Different colors based on breakpoint size
                                    if bp_px <= 480:  # Mobile
                                        bg_color = "FFF8E1"  # Light amber for mobile
                                    elif bp_px <= 768:  # Tablet
                                        bg_color = "FFECB3"  # Medium amber for tablet
                                    else:  # Desktop
                                        bg_color = "FFE0B2"  # Dark amber for desktop
                                    
                                    cell.fill = openpyxl.styles.PatternFill(
                                        start_color=bg_color,
                                        end_color=bg_color,
                                        fill_type="solid"
                                    )
                                except:
                                    # Default formatting if parsing fails
                                    cell.fill = openpyxl.styles.PatternFill(
                                        start_color="F5F5F5",
                                        end_color="F5F5F5",
                                        fill_type="solid"
                                    )
                            
                            # Check for the Total Issues column
                            header_cell = worksheet.cell(row=1, column=i+1)
                            if header_cell.value == 'Total Issues':
                                total_issues_col = i
                        
                        # Format the Total Issues column based on value
                        if total_issues_col is not None and row[total_issues_col].value:
                            try:
                                issue_count = int(row[total_issues_col].value)
                                if issue_count > 5:
                                    row[total_issues_col].fill = openpyxl.styles.PatternFill(
                                        start_color="FFCCCC",  # Light red (high)
                                        end_color="FFCCCC",
                                        fill_type="solid"
                                    )
                                    row[total_issues_col].font = openpyxl.styles.Font(size=16, bold=True)
                                elif issue_count > 0:
                                    row[total_issues_col].fill = openpyxl.styles.PatternFill(
                                        start_color="FFE0B2",  # Light orange (medium)
                                        end_color="FFE0B2",
                                        fill_type="solid"
                                    )
                            except:
                                pass
                
                # Format data rows (general formatting for all sheets)
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if not cell.font.size:  # Only set font if not already set
                            cell.font = openpyxl.styles.Font(size=16)
                        
                        # Check if content is a table-like structure
                        if isinstance(cell.value, str) and '\n' in cell.value:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=True,
                                vertical='top',
                                horizontal='left'
                            )
                            # Calculate row height based on content
                            line_count = cell.value.count('\n') + 1
                            min_height = 20 * line_count  # 20 pixels per line
                            current_height = worksheet.row_dimensions[cell.row].height or 15
                            worksheet.row_dimensions[cell.row].height = max(min_height, current_height)
                        else:
                            # Regular cell formatting
                            if cell.column == 1:  # Column A
                                cell.alignment = openpyxl.styles.Alignment(
                                    wrap_text=True,
                                    vertical='top',
                                    horizontal='right'
                                )
                            else:
                                cell.alignment = openpyxl.styles.Alignment(
                                    wrap_text=True,
                                    vertical='top',
                                    horizontal='left'
                                )
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if cell.value:
                                # For table-like content, consider the longest line
                                if isinstance(cell.value, str) and '\n' in cell.value:
                                    lines = cell.value.split('\n')
                                    max_length = max(len(line) for line in lines)
                                else:
                                    max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

def main(db_name=None):
    try:
        # Initialize database connection
        db = AccessibilityDB(db_name=db_name)
        
        # Analyze database structure
        print("Analyzing database structure...")
        analyzer = TemplateAnalyzer(db)
        structures, examples = analyzer.analyze_test_structures()
        analyzer.print_analysis()
        
        # Generate report
        print("\nGenerating accessibility report...")
        generator = AccessibilityReportGenerator(db, structures, examples)
        
        # Get regular test run IDs
        test_run_ids = db.get_all_test_run_ids()
        
        # Also add documentation test run if it exists
        doc_test_run = db.get_test_run_by_name("Documentation Test Run")
        if doc_test_run:
            doc_test_run_id = str(doc_test_run['_id'])
            print(f"Including Documentation Test Run: {doc_test_run_id}")
            # Add documentation test run ID if not already in list
            if doc_test_run_id not in test_run_ids:
                test_run_ids.append(doc_test_run_id)
                
            # Directly get the example.com test result and process its documentation
            example_result = db.db.page_results.find_one({"url": "https://example.com"})
            if example_result:
                print("Found example.com test result - processing documentation directly")
                # Process the documentation in advance
                accessibility_tests = example_result.get('results', {}).get('accessibility', {}).get('tests', {})
                if 'page_structure' in accessibility_tests and 'page_structure' in accessibility_tests['page_structure']:
                    doc = accessibility_tests['page_structure']['page_structure'].get('documentation')
                    if doc:
                        print(f"Pre-loading Page Structure documentation with {len(doc.get('tests', []))} individual tests")
                        generator.test_documentation['page_structure'] = doc
                
                if 'accessible_names' in accessibility_tests and 'accessible_names' in accessibility_tests['accessible_names']:
                    doc = accessibility_tests['accessible_names']['accessible_names'].get('documentation')
                    if doc:
                        print(f"Pre-loading Accessible Names documentation with {len(doc.get('tests', []))} individual tests")
                        generator.test_documentation['accessible_names'] = doc
                
                if 'focus_management' in accessibility_tests:
                    # Debug output to inspect the focus_management data structure
                    print("  Found focus_management in tests, examining structure...")
                    fm_data = accessibility_tests['focus_management']
                    
                    # Option 1: Nested structure with focus_management.focus_management.documentation
                    if isinstance(fm_data, dict) and 'focus_management' in fm_data and isinstance(fm_data['focus_management'], dict):
                        print("  Checking for nested focus_management > focus_management > documentation structure")
                        if 'documentation' in fm_data['focus_management']:
                            doc = fm_data['focus_management']['documentation']
                            print(f"  Found nested documentation with {len(doc.get('tests', []))} individual tests")
                            generator.test_documentation['focus_management'] = doc
                    
                    # Option 2: Direct documentation in focus_management
                    elif isinstance(fm_data, dict) and 'documentation' in fm_data:
                        print("  Checking for direct focus_management > documentation structure")
                        doc = fm_data['documentation']
                        print(f"  Found direct documentation with {len(doc.get('tests', []))} individual tests")
                        generator.test_documentation['focus_management'] = doc
                    
                    # Option 3: Focus management as direct document  
                    elif isinstance(fm_data, dict):
                        print("  Dumping focus_management keys for diagnosis: " + str(list(fm_data.keys())))
        
        if test_run_ids:
            print(f"Generating report for {len(test_run_ids)} test runs")
            output_file = 'accessibility_report.xlsx'
            generator.generate_excel_report(test_run_ids, output_file, db_name=db.db_name)
            
            # Determine the actual filename after possible db_name prepending
            if db.db_name != DEFAULT_DB_NAME:
                base, ext = output_file.rsplit('.', 1)
                actual_file = f"{db.db_name}_{base}.{ext}"
            else:
                actual_file = output_file
                
            print(f"Report generated successfully: {actual_file}")
        else:
            print("No test runs found in the database")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Generate Excel report from MongoDB accessibility test results')
    parser.add_argument('--database', '-db', help='MongoDB database name to use (default: accessibility_tests)')
    args = parser.parse_args()
    
    main(db_name=args.database)